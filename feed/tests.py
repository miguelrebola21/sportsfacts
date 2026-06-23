from io import BytesIO
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook, load_workbook
from PIL import Image

from .models import Fact, Record, Tag
from .xlsx import XlsxImportError, export_global_xlsx, export_xlsx, import_global_xlsx, import_xlsx


class FeedModelTests(TestCase):
    def test_facts_and_records_can_share_tags(self):
        tag = Tag.objects.create(name="basketball")
        fact = Fact.objects.create(text="the first basketballs were brown")
        record = Record.objects.create(number=232, text="most points in a game")

        fact.tags.add(tag)
        record.tags.add(tag)

        self.assertEqual(list(fact.tags.values_list("name", flat=True)), ["basketball"])
        self.assertEqual(list(record.tags.values_list("name", flat=True)), ["basketball"])

    def test_facts_and_records_can_invalidate_previous_items(self):
        old_fact = Fact.objects.create(text="old fact")
        new_fact = Fact.objects.create(text="new fact", invalidates=old_fact)
        old_record = Record.objects.create(number=232, text="old record")
        new_record = Record.objects.create(number=233, text="new record", invalidates=old_record)

        self.assertEqual(new_fact.invalidates, old_fact)
        self.assertEqual(new_record.invalidates, old_record)


class FeedViewTests(TestCase):
    def test_home_page_lists_fact_and_record_cards(self):
        Fact.objects.create(text="Blgvaskjfadksfa")
        Record.objects.create(number=232, text="dlkaflasdklfkdasf")

        response = self.client.get(reverse("feed:home"))

        self.assertContains(response, "Fact #1")
        self.assertContains(response, "Blgvaskjfadksfa")
        self.assertContains(response, "Record #232")
        self.assertContains(response, "dlkaflasdklfkdasf")
        self.assertContains(response, "#002030")
        self.assertContains(response, "feed/logo.svg")
        self.assertContains(response, "max-width: 580px")
        self.assertContains(response, "card fact-card")
        self.assertContains(response, "card record-card")
        self.assertContains(response, 'class="scroll-top"')
        self.assertContains(response, 'aria-label="scroll to top"')

    def test_home_page_links_invalidated_items(self):
        old_fact = Fact.objects.create(text="old fact")
        Fact.objects.create(text="new fact", invalidates=old_fact)
        old_record = Record.objects.create(number=232, text="old record")
        Record.objects.create(number=233, text="new record", invalidates=old_record)

        response = self.client.get(reverse("feed:home"))

        self.assertContains(response, 'href="#fact-1"')
        self.assertContains(response, "<s>Fact #1")
        self.assertContains(response, "old fact")
        self.assertContains(response, 'href="#record-1"')
        self.assertContains(response, "<s>Record #232")
        self.assertContains(response, "old record")

    def test_home_page_scrambles_fact_and_record_cards(self):
        Fact.objects.create(text="fact")
        Fact.objects.create(text="second fact")
        Record.objects.create(number=232, text="record")
        Record.objects.create(number=233, text="second record")

        with patch("feed.views.random.shuffle", side_effect=lambda cards: None):
            response = self.client.get(reverse("feed:home"))

        content = response.content.decode()
        positions = [
            content.index("Fact #1"),
            content.index("Record #232"),
            content.index("Fact #2"),
            content.index("Record #233"),
        ]
        self.assertEqual(positions, sorted(positions))

    def test_home_page_shows_vote_and_share_controls(self):
        Fact.objects.create(text="hgHGFADHGFHS")
        Record.objects.create(number=232, text="dlkaflasdklfkdasf")

        response = self.client.get(reverse("feed:home"))

        self.assertContains(response, 'aria-label="upvote"')
        self.assertContains(response, 'aria-label="downvote"')
        self.assertContains(response, 'aria-label="share on x"')
        self.assertContains(response, 'aria-label="share on facebook"')
        self.assertContains(response, 'aria-label="share on linkedin"')
        self.assertContains(response, 'aria-label="share on whatsapp"')
        self.assertContains(response, "share%2Ffact%2F1%2F")
        self.assertContains(response, "<svg")

    def test_fact_upvote_and_downvote_increment_counts(self):
        fact = Fact.objects.create(text="hgHGFADHGFHS")

        self.client.post(reverse("feed:vote", args=["fact", fact.id, "up"]))
        self.client.post(reverse("feed:vote", args=["fact", fact.id, "down"]))

        fact.refresh_from_db()
        self.assertEqual(fact.upvotes, 1)
        self.assertEqual(fact.downvotes, 1)

    def test_record_upvote_and_downvote_increment_counts(self):
        record = Record.objects.create(number=232, text="dlkaflasdklfkdasf")

        self.client.post(reverse("feed:vote", args=["record", record.id, "up"]))
        self.client.post(reverse("feed:vote", args=["record", record.id, "down"]))

        record.refresh_from_db()
        self.assertEqual(record.upvotes, 1)
        self.assertEqual(record.downvotes, 1)

    def test_share_image_returns_png(self):
        fact = Fact.objects.create(text="hgHGFADHGFHS")

        response = self.client.get(reverse("feed:share_image", args=["fact", fact.id]))

        self.assertEqual(response["Content-Type"], "image/png")
        self.assertTrue(response.content.startswith(b"\x89PNG"))
        image = Image.open(BytesIO(response.content))
        self.assertEqual(image.size, (1200, 630))
        self.assertEqual(image.getpixel((0, 0)), (0, 32, 48))

    def test_share_page_has_social_image_metadata(self):
        fact = Fact.objects.create(text="share fact")

        response = self.client.get(reverse("feed:share", args=["fact", fact.id]))

        self.assertContains(response, '<meta property="og:title" content="Fact #1">')
        self.assertContains(response, '<meta property="og:image" content="http://testserver/share/fact/1.png">')
        self.assertContains(response, '<meta name="twitter:card" content="summary_large_image">')

    def test_card_tags_add_multiple_selected_filters(self):
        basketball = Tag.objects.create(name="basketball")
        football = Tag.objects.create(name="football")
        tennis = Tag.objects.create(name="tennis")
        fact = Fact.objects.create(text="basketball fact")
        record = Record.objects.create(number=232, text="football record")
        multi_tag_fact = Fact.objects.create(text="basketball football fact")
        hidden_fact = Fact.objects.create(text="tennis fact")
        fact.tags.add(basketball)
        record.tags.add(football)
        multi_tag_fact.tags.add(basketball, football)
        hidden_fact.tags.add(tennis)

        response = self.client.get(reverse("feed:home"), {"tags": ["basketball", "football"]})

        self.assertContains(response, "basketball")
        self.assertContains(response, "football")
        self.assertContains(response, "×")
        self.assertNotContains(response, "min-width: 100px")
        self.assertNotContains(response, ">all</a>")
        self.assertNotContains(response, "selected tags")
        self.assertNotContains(response, 'class="selected-tags"')
        self.assertContains(response, "basketball")
        self.assertContains(response, "football")
        self.assertContains(response, "?tags=football")
        self.assertContains(response, "?tags=basketball")
        self.assertContains(response, "basketball football fact")
        self.assertNotContains(response, "basketball fact")
        self.assertNotContains(response, "football record")
        self.assertNotContains(response, "tennis fact")

    def test_card_tag_links_add_to_current_selection(self):
        basketball = Tag.objects.create(name="basketball")
        football = Tag.objects.create(name="football")
        fact = Fact.objects.create(text="multi tag fact")
        fact.tags.add(basketball, football)

        response = self.client.get(reverse("feed:home"), {"tags": ["basketball"]})

        self.assertContains(response, "basketball")
        self.assertContains(response, "×")
        self.assertContains(response, "?tags=basketball&amp;tags=football")


class SeedFixtureTests(TestCase):
    fixtures = ["seed.json"]

    def test_seed_fixture_loads_feed_content(self):
        self.assertTrue(Fact.objects.filter(text="Blgvaskjfadksfa").exists())
        self.assertTrue(Record.objects.filter(number=232, text="dlkaflasdklfkdasf").exists())
        self.assertEqual(Tag.objects.count(), 2)


class XlsxImportExportTests(TestCase):
    def test_exports_facts_with_tags(self):
        tag = Tag.objects.create(name="basketball")
        old_fact = Fact.objects.create(text="old fact")
        fact = Fact.objects.create(text="exported fact", invalidates=old_fact, upvotes=2, downvotes=1)
        fact.tags.add(tag)

        content = export_xlsx(Fact, Fact.objects.all())
        workbook = load_workbook(BytesIO(content))
        rows = list(workbook.active.iter_rows(values_only=True))

        self.assertEqual(rows[0], ("id", "text", "tags", "invalidates_id", "upvotes", "downvotes"))
        self.assertEqual(rows[2], (fact.id, "exported fact", "basketball", old_fact.id, 2, 1))

    def test_imports_facts_and_creates_tags(self):
        workbook = Workbook()
        sheet = workbook.active
        old_fact = Fact.objects.create(text="old fact")
        sheet.append(["id", "text", "tags", "invalidates_id", "upvotes", "downvotes"])
        sheet.append([None, "imported fact", "basketball, football", old_fact.id, 3, 1])
        content = BytesIO()
        workbook.save(content)

        result = import_xlsx(Fact, BytesIO(content.getvalue()))

        fact = Fact.objects.get(text="imported fact")
        self.assertEqual(result, 1)
        self.assertEqual(fact.upvotes, 3)
        self.assertEqual(fact.downvotes, 1)
        self.assertEqual(fact.invalidates, old_fact)
        self.assertEqual(set(fact.tags.values_list("name", flat=True)), {"basketball", "football"})

    def test_imports_records_and_tags(self):
        workbook = Workbook()
        sheet = workbook.active
        old_record = Record.objects.create(number=231, text="old record")
        sheet.append(["id", "number", "text", "tags", "invalidates_id", "upvotes", "downvotes"])
        sheet.append([None, 232, "imported record", "football", old_record.id, 4, 2])
        content = BytesIO()
        workbook.save(content)

        result = import_xlsx(Record, BytesIO(content.getvalue()))

        record = Record.objects.get(number=232)
        self.assertEqual(result, 1)
        self.assertEqual(record.text, "imported record")
        self.assertEqual(record.upvotes, 4)
        self.assertEqual(record.downvotes, 2)
        self.assertEqual(record.invalidates, old_record)
        self.assertEqual(list(record.tags.values_list("name", flat=True)), ["football"])

    def test_exports_global_workbook(self):
        tag = Tag.objects.create(name="basketball")
        fact = Fact.objects.create(text="global fact")
        record = Record.objects.create(number=232, text="global record")
        fact.tags.add(tag)
        record.tags.add(tag)

        workbook = load_workbook(BytesIO(export_global_xlsx()))

        self.assertEqual(workbook.sheetnames, ["Tags", "Facts", "Records"])
        self.assertEqual(workbook["Tags"]["A1"].value, "name")
        self.assertEqual(workbook["Facts"]["B2"].value, "global fact")
        self.assertEqual(workbook["Records"]["C2"].value, "global record")

    def test_imports_global_workbook(self):
        workbook = Workbook()
        tags_sheet = workbook.active
        tags_sheet.title = "Tags"
        tags_sheet.append(["name"])
        tags_sheet.append(["basketball"])
        facts_sheet = workbook.create_sheet("Facts")
        facts_sheet.append(["id", "text", "tags", "invalidates_id", "upvotes", "downvotes"])
        facts_sheet.append([None, "global imported fact", "basketball", None, 1, 0])
        records_sheet = workbook.create_sheet("Records")
        records_sheet.append(["id", "number", "text", "tags", "invalidates_id", "upvotes", "downvotes"])
        records_sheet.append([None, 232, "global imported record", "basketball", None, 2, 1])
        content = BytesIO()
        workbook.save(content)

        result = import_global_xlsx(BytesIO(content.getvalue()))

        self.assertEqual(result, {"tags": 1, "facts": 1, "records": 1})
        self.assertTrue(Fact.objects.filter(text="global imported fact", tags__name="basketball").exists())
        self.assertTrue(Record.objects.filter(number=232, tags__name="basketball").exists())

    def test_global_import_validates_sheet_headers(self):
        workbook = Workbook()
        records_sheet = workbook.active
        records_sheet.title = "Records"
        records_sheet.append(["id", "text", "tags", "upvotes", "downvotes"])
        records_sheet.append([None, "Usain Bolt ran the men's 100 m world record of 9.58 seconds in 2009.", "track", 0, 0])
        content = BytesIO()
        workbook.save(content)

        with self.assertRaisesMessage(XlsxImportError, "Records sheet must have columns"):
            import_global_xlsx(BytesIO(content.getvalue()))


class XlsxAdminTests(TestCase):
    def setUp(self):
        user = get_user_model().objects.create_superuser("admin", "admin@example.com", "password")
        self.client.force_login(user)

    def test_admin_changelist_has_import_and_export_links(self):
        response = self.client.get(reverse("admin:feed_fact_changelist"))

        self.assertContains(response, "import xlsx")
        self.assertContains(response, "export xlsx")
        self.assertContains(response, "global import xlsx")
        self.assertContains(response, "global export xlsx")
        self.assertContains(response, "import-xlsx/")
        self.assertContains(response, "export-xlsx/")
        self.assertContains(response, "/admin/feed/import-xlsx/")
        self.assertContains(response, "/admin/feed/export-xlsx/")

    def test_admin_export_downloads_xlsx(self):
        Fact.objects.create(text="admin export")

        response = self.client.get(reverse("admin:feed_fact_export_xlsx"))

        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertTrue(response.content.startswith(b"PK"))

    def test_admin_global_export_downloads_xlsx(self):
        response = self.client.get(reverse("admin:feed_global_export_xlsx"))

        self.assertEqual(response["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.assertTrue(response.content.startswith(b"PK"))
