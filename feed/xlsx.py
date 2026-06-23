from io import BytesIO

from openpyxl import Workbook, load_workbook

from .models import Fact, Record, Tag

HEADERS = {
    Tag: ("name",),
    Fact: ("id", "text", "tags", "invalidates_id", "upvotes", "downvotes"),
    Record: ("id", "number", "text", "tags", "invalidates_id", "upvotes", "downvotes"),
}


class XlsxImportError(ValueError):
    pass


def export_global_xlsx():
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _add_model_sheet(workbook, Tag, Tag.objects.all())
    _add_model_sheet(workbook, Fact, Fact.objects.prefetch_related("tags"))
    _add_model_sheet(workbook, Record, Record.objects.prefetch_related("tags"))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def import_global_xlsx(file_obj):
    workbook = load_workbook(file_obj)
    return {
        "tags": _import_sheet(Tag, workbook["Tags"]) if "Tags" in workbook.sheetnames else 0,
        "facts": _import_sheet(Fact, workbook["Facts"]) if "Facts" in workbook.sheetnames else 0,
        "records": _import_sheet(Record, workbook["Records"]) if "Records" in workbook.sheetnames else 0,
    }


def export_xlsx(model, queryset):
    workbook = Workbook()
    sheet = workbook.active
    _write_model_sheet(sheet, model, queryset)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def import_xlsx(model, file_obj):
    workbook = load_workbook(file_obj)
    return _import_sheet(model, workbook.active)


def _add_model_sheet(workbook, model, queryset):
    sheet = workbook.create_sheet(model._meta.verbose_name_plural.title())
    _write_model_sheet(sheet, model, queryset)


def _write_model_sheet(sheet, model, queryset):
    sheet.title = model._meta.verbose_name_plural.title()
    sheet.append(HEADERS[model])

    for item in queryset:
        if model is Tag:
            sheet.append([item.name])
        elif model is Fact:
            sheet.append([item.id, item.text, _tag_names(item), item.invalidates_id, item.upvotes, item.downvotes])
        elif model is Record:
            sheet.append([item.id, item.number, item.text, _tag_names(item), item.invalidates_id, item.upvotes, item.downvotes])


def _import_sheet(model, sheet):
    headers = [str(value).strip().lower() for value in next(sheet.iter_rows(values_only=True))]
    expected_headers = list(HEADERS[model])
    if headers != expected_headers:
        raise XlsxImportError(f"{sheet.title} sheet must have columns: {', '.join(expected_headers)}")
    imported = 0

    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if _is_empty_row(row):
            continue
        _import_row(model, row)
        imported += 1

    return imported


def _import_row(model, row):
    if model is Tag:
        name = _text(row.get("name"))
        if name:
            Tag.objects.get_or_create(name=name)
        return

    if model is Fact:
        item_id = _int_or_none(row.get("id"))
        defaults = {
            "text": _text(row.get("text")),
            "invalidates_id": _validated_invalidates_id(Fact, row.get("invalidates_id")),
            "upvotes": _int_or_zero(row.get("upvotes")),
            "downvotes": _int_or_zero(row.get("downvotes")),
        }
        item = _update_or_create(Fact, item_id, defaults)
        item.tags.set(_get_tags(row.get("tags")))
        return

    if model is Record:
        item_id = _int_or_none(row.get("id"))
        defaults = {
            "number": _int_or_zero(row.get("number")),
            "text": _text(row.get("text")),
            "invalidates_id": _validated_invalidates_id(Record, row.get("invalidates_id")),
            "upvotes": _int_or_zero(row.get("upvotes")),
            "downvotes": _int_or_zero(row.get("downvotes")),
        }
        item = _update_or_create(Record, item_id, defaults)
        item.tags.set(_get_tags(row.get("tags")))


def _update_or_create(model, item_id, defaults):
    if item_id:
        item, _ = model.objects.update_or_create(id=item_id, defaults=defaults)
        return item
    return model.objects.create(**defaults)


def _get_tags(value):
    tags = []
    for name in _tag_values(value):
        tag, _ = Tag.objects.get_or_create(name=name)
        tags.append(tag)
    return tags


def _validated_invalidates_id(model, value):
    item_id = _int_or_none(value)
    if item_id and not model.objects.filter(id=item_id).exists():
        raise XlsxImportError(f"Invalid invalidates_id {item_id} for {model._meta.verbose_name}.")
    return item_id


def _tag_names(item):
    return ", ".join(item.tags.values_list("name", flat=True))


def _tag_values(value):
    return [name.strip() for name in _text(value).split(",") if name and name.strip()]


def _is_empty_row(row):
    return all(value is None or value == "" for value in row.values())


def _int_or_none(value):
    if value in (None, ""):
        return None
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise XlsxImportError(f"Expected an integer, got {value!r}") from exc


def _int_or_zero(value):
    if value in (None, ""):
        return 0
    try:
        return int(value)
    except (TypeError, ValueError) as exc:
        raise XlsxImportError(f"Expected an integer, got {value!r}") from exc


def _text(value):
    if value is None:
        return ""
    return str(value).strip()
